# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'funcionarios.xlsx'

workbook: Workbook = Workbook()
workbook.remove_sheet(workbook['Sheet'])
for i in range(1):
    sheet_name = f'FUNC{i+1}'
    workbook.create_sheet(sheet_name)
    worksheet: Worksheet = workbook[sheet_name]

    # Remover sheet

    # Criando os cabeçalhos
    worksheet.cell(1, 1, 'Nome Empresa')
    worksheet.cell(1, 2, 'Nome')
    worksheet.cell(1, 3, 'Exames')

    # students = [
    #     # nome      idade nota
    #     ['João',    14,   5.5],
    #     ['Maria',   13,   9.7],
    #     ['Luiz',    15,   8.8],
    #     ['Alberto', 16,   10],
    # ]

    students = [
        # nome Empresa Nome funcionario     exames feitos    nota
        ['Empresa1',       'João',              'teste1/exame1'],
        ['Empresa2',       'Maria',             'teste2/exame2'],
        ['Empresa3',       'Luiz',              'teste3/exame3'],
        ['Empresa1',       'Alberto',           'teste4/exame4'],
        ['Empresa1',       'Matheus',           'teste3/exame1'],
        ['Empresa2',       'Luiz',           'teste1/exame4'],
        ['Empresa2',       'Samuel',           'teste4/exame1'],
        ['Empresa3',       'Edna',           'teste2/exame2'],
        ['Empresa3',       'Lucas',           'teste3/exame3'],
        ['Empresa1',       'Fernanda',           'teste4/exame1'],
        ['Empresa2',      'Cristiano',           'teste2/exame4'],
    ]
    # for i, student_row in enumerate(students, start=2):
    #     for j, student_column in enumerate(student_row, start=1):
    #         worksheet.cell(i, j, student_column)

    for student in students:
        worksheet.append(student)

workbook.save(WORKBOOK_PATH)
